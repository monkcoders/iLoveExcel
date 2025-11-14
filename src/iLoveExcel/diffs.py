"""
iLoveExcel - CSV Side-by-Side Diff module.

This module provides functions to compare two CSV/Excel files side-by-side
and generate diff reports with highlighting.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .io import read_csv_chunked

logger = logging.getLogger(__name__)

# Color definitions for diff highlighting
COLORS = {
    'match': 'C6EFCE',      # Light green
    'diff': 'FFEB9C',       # Light yellow
    'only_a': 'BDD7EE',     # Light blue
    'only_b': 'F8CBAD',     # Light red/orange
    'header': 'D9D9D9',     # Light gray
}


def diff_csv_side_by_side(
    file_a: Union[str, Path],
    file_b: Union[str, Path],
    key_columns: Optional[List[str]] = None,
    compare_by_index: bool = True,
    ignore_whitespace: bool = False,
    case_insensitive: bool = False,
    ignore_column_order: bool = False,
    show_only_diffs: bool = False,
    max_rows: Optional[int] = None,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    Compare two CSV/Excel files side-by-side.
    
    Args:
        file_a: Path to first file (left)
        file_b: Path to second file (right)
        key_columns: List of column names to use as key for alignment (if compare_by_index=False)
        compare_by_index: If True, compare by row index; if False, use key_columns
        ignore_whitespace: Strip whitespace before comparison
        case_insensitive: Ignore case when comparing strings
        ignore_column_order: Match columns by name regardless of order
        show_only_diffs: Return only rows with differences
        max_rows: Maximum number of rows to process (None = unlimited)
    
    Returns:
        Tuple of (diff_dataframe, stats_dict)
        - diff_dataframe: DataFrame with columns suffixed _A and _B, plus 'Status' column
        - stats_dict: Dictionary with 'total', 'matching', 'different', 'only_a', 'only_b'
    
    Raises:
        FileNotFoundError: If either file doesn't exist
        ValueError: If key_columns not found or files incompatible
    """
    file_a = Path(file_a)
    file_b = Path(file_b)
    
    if not file_a.exists():
        raise FileNotFoundError(f"File A not found: {file_a}")
    if not file_b.exists():
        raise FileNotFoundError(f"File B not found: {file_b}")
    
    logger.info(f"Comparing {file_a} vs {file_b}")
    
    # Read files
    df_a = _read_file(file_a, max_rows)
    df_b = _read_file(file_b, max_rows)
    
    logger.info(f"File A: {len(df_a)} rows, {len(df_a.columns)} columns")
    logger.info(f"File B: {len(df_b)} rows, {len(df_b.columns)} columns")
    
    # Preprocess data
    if ignore_whitespace:
        df_a = _strip_whitespace(df_a)
        df_b = _strip_whitespace(df_b)
    
    if case_insensitive:
        df_a = _lowercase_strings(df_a)
        df_b = _lowercase_strings(df_b)
    
    # Align dataframes
    if compare_by_index:
        df_a_aligned, df_b_aligned = _align_by_index(df_a, df_b, ignore_column_order)
    else:
        if not key_columns:
            raise ValueError("key_columns must be provided when compare_by_index=False")
        df_a_aligned, df_b_aligned = _align_by_key(df_a, df_b, key_columns, ignore_column_order)
    
    # Perform comparison
    diff_df, stats = _compare_dataframes(df_a_aligned, df_b_aligned, show_only_diffs)
    
    logger.info(f"Comparison stats: {stats}")
    return diff_df, stats


def export_diff_to_excel(
    diff_df: pd.DataFrame,
    stats: Dict[str, int],
    output_path: Union[str, Path],
    file_a_name: str = "File A",
    file_b_name: str = "File B",
    highlight: bool = True,
) -> None:
    """
    Export diff results to Excel with optional highlighting.
    
    Creates a multi-sheet workbook:
    - Sheet 1: Side-by-side comparison with highlighting
    - Sheet 2: Summary statistics
    - Sheet 3: Rows only in A
    - Sheet 4: Rows only in B
    
    Args:
        diff_df: DataFrame returned by diff_csv_side_by_side()
        stats: Stats dictionary from diff_csv_side_by_side()
        output_path: Path to output Excel file
        file_a_name: Display name for file A
        file_b_name: Display name for file B
        highlight: Whether to apply color highlighting
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"Exporting diff to Excel: {output_path}")
    
    wb = Workbook()
    
    # Sheet 1: Side-by-side comparison
    ws_main = wb.active
    ws_main.title = "Comparison"
    
    if highlight:
        _write_comparison_sheet_with_highlights(ws_main, diff_df, file_a_name, file_b_name)
    else:
        _write_comparison_sheet_plain(ws_main, diff_df)
    
    # Sheet 2: Summary statistics
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, stats, file_a_name, file_b_name)
    
    # Sheet 3: Rows only in A
    ws_only_a = wb.create_sheet(f"Only in {file_a_name}")
    only_a_df = diff_df[diff_df['Status'] == 'ONLY_A']
    _write_dataframe_to_sheet(ws_only_a, only_a_df)
    
    # Sheet 4: Rows only in B
    ws_only_b = wb.create_sheet(f"Only in {file_b_name}")
    only_b_df = diff_df[diff_df['Status'] == 'ONLY_B']
    _write_dataframe_to_sheet(ws_only_b, only_b_df)
    
    wb.save(output_path)
    logger.info(f"Exported diff to {output_path}")


# ============================================================================
# Helper Functions
# ============================================================================

def _read_file(file_path: Path, max_rows: Optional[int] = None) -> pd.DataFrame:
    """Read CSV or Excel file."""
    if file_path.suffix.lower() == '.csv':
        df = read_csv_chunked(file_path, chunksize=None)
    else:  # Excel
        df = pd.read_excel(file_path)
    
    if max_rows is not None and len(df) > max_rows:
        logger.warning(f"Truncating {file_path} from {len(df)} to {max_rows} rows")
        df = df.head(max_rows)
    
    return df


def _strip_whitespace(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from string columns."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
    return df


def _lowercase_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Convert string columns to lowercase."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.lower()
    return df


def _align_by_index(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    ignore_column_order: bool
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Align dataframes by row index."""
    # Align columns
    if ignore_column_order:
        common_cols = sorted(set(df_a.columns) & set(df_b.columns))
        df_a = df_a[common_cols]
        df_b = df_b[common_cols]
    else:
        # Use columns from A, fill missing with NaN
        all_cols = list(df_a.columns)
        for col in df_b.columns:
            if col not in all_cols:
                all_cols.append(col)
        
        df_a = df_a.reindex(columns=all_cols)
        df_b = df_b.reindex(columns=all_cols)
    
    # Align rows by index
    all_indices = sorted(set(df_a.index) | set(df_b.index))
    df_a = df_a.reindex(all_indices)
    df_b = df_b.reindex(all_indices)
    
    return df_a, df_b


def _align_by_key(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    key_columns: List[str],
    ignore_column_order: bool
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Align dataframes by key columns."""
    # Validate key columns
    for col in key_columns:
        if col not in df_a.columns:
            raise ValueError(f"Key column '{col}' not found in file A")
        if col not in df_b.columns:
            raise ValueError(f"Key column '{col}' not found in file B")
    
    # Perform outer merge on key columns
    df_a_keyed = df_a.copy()
    df_b_keyed = df_b.copy()
    
    # Add row number for tracking
    df_a_keyed['_row_a'] = range(len(df_a))
    df_b_keyed['_row_b'] = range(len(df_b))
    
    merged = pd.merge(
        df_a_keyed,
        df_b_keyed,
        on=key_columns,
        how='outer',
        suffixes=('_a', '_b'),
        indicator=True
    )
    
    # Separate back into A and B
    df_a_aligned = merged[[c for c in merged.columns if c.endswith('_a') or c in key_columns or c == '_merge']]
    df_b_aligned = merged[[c for c in merged.columns if c.endswith('_b') or c in key_columns or c == '_merge']]
    
    # Clean up column names
    df_a_aligned.columns = [c.replace('_a', '') for c in df_a_aligned.columns]
    df_b_aligned.columns = [c.replace('_b', '') for c in df_b_aligned.columns]
    
    return df_a_aligned, df_b_aligned


def _compare_dataframes(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    show_only_diffs: bool
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Compare two aligned dataframes and return diff results."""
    # Create result dataframe
    result_rows = []
    
    stats = {
        'total': 0,
        'matching': 0,
        'different': 0,
        'only_a': 0,
        'only_b': 0,
    }
    
    for idx in df_a.index:
        row_a = df_a.loc[idx]
        row_b = df_b.loc[idx]
        
        # Determine row status
        is_only_a = row_a.notna().any() and row_b.isna().all()
        is_only_b = row_a.isna().all() and row_b.notna().any()
        
        if is_only_a:
            status = 'ONLY_A'
            stats['only_a'] += 1
        elif is_only_b:
            status = 'ONLY_B'
            stats['only_b'] += 1
        else:
            # Compare values
            has_diff = False
            for col in df_a.columns:
                val_a = row_a.get(col)
                val_b = row_b.get(col)
                
                # Handle NaN comparison
                if pd.isna(val_a) and pd.isna(val_b):
                    continue
                elif val_a != val_b:
                    has_diff = True
                    break
            
            if has_diff:
                status = 'DIFF'
                stats['different'] += 1
            else:
                status = 'MATCH'
                stats['matching'] += 1
        
        stats['total'] += 1
        
        # Add to results if not filtering or if has difference
        if not show_only_diffs or status != 'MATCH':
            result_row = {
                'Row_Index': idx,
                'Status': status,
            }
            
            # Add A and B columns
            for col in df_a.columns:
                result_row[f'{col}_A'] = row_a.get(col)
                result_row[f'{col}_B'] = row_b.get(col)
            
            result_rows.append(result_row)
    
    result_df = pd.DataFrame(result_rows)
    return result_df, stats


def _write_comparison_sheet_with_highlights(
    ws,
    diff_df: pd.DataFrame,
    file_a_name: str,
    file_b_name: str
) -> None:
    """Write comparison sheet with color highlighting."""
    # Define fills
    fill_match = PatternFill(start_color=COLORS['match'], end_color=COLORS['match'], fill_type='solid')
    fill_diff = PatternFill(start_color=COLORS['diff'], end_color=COLORS['diff'], fill_type='solid')
    fill_only_a = PatternFill(start_color=COLORS['only_a'], end_color=COLORS['only_a'], fill_type='solid')
    fill_only_b = PatternFill(start_color=COLORS['only_b'], end_color=COLORS['only_b'], fill_type='solid')
    fill_header = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    
    font_header = Font(bold=True)
    
    # Write headers
    headers = list(diff_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill_header
        cell.font = font_header
    
    # Write data with highlighting
    for row_idx, row in enumerate(diff_df.itertuples(index=False), start=2):
        status = row.Status
        
        # Determine fill based on status
        if status == 'MATCH':
            fill = fill_match
        elif status == 'DIFF':
            fill = fill_diff
        elif status == 'ONLY_A':
            fill = fill_only_a
        elif status == 'ONLY_B':
            fill = fill_only_b
        else:
            fill = None
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill


def _write_comparison_sheet_plain(ws, diff_df: pd.DataFrame) -> None:
    """Write comparison sheet without highlighting."""
    for r_idx, row in enumerate(dataframe_to_rows(diff_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _write_summary_sheet(
    ws,
    stats: Dict[str, int],
    file_a_name: str,
    file_b_name: str
) -> None:
    """Write summary statistics sheet."""
    ws.append(['Comparison Summary'])
    ws.append([])
    ws.append([f'File A: {file_a_name}'])
    ws.append([f'File B: {file_b_name}'])
    ws.append([])
    ws.append(['Statistic', 'Count'])
    ws.append(['Total Rows Compared', stats['total']])
    ws.append(['Matching Rows', stats['matching']])
    ws.append(['Different Rows', stats['different']])
    ws.append([f'Rows Only in {file_a_name}', stats['only_a']])
    ws.append([f'Rows Only in {file_b_name}', stats['only_b']])
    
    # Bold first column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)


def _write_dataframe_to_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)
