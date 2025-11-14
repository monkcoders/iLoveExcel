"""
iLoveExcel - I/O Helper Functions.

Additional I/O utilities including auto-column-width for Excel files.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Default configuration for auto-width
DEFAULT_AUTO_WIDTH_CONFIG = {
    'min_width': 8,      # Minimum column width
    'max_width': 50,     # Maximum column width
    'padding': 2,        # Extra padding (characters)
    'header_factor': 1.2,  # Multiply header length by this factor
}


def apply_auto_column_width(
    excel_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    min_width: Optional[int] = None,
    max_width: Optional[int] = None,
    padding: Optional[int] = None,
    header_factor: Optional[float] = None,
) -> None:
    """
    Apply auto-width to columns in an Excel file.
    
    Adjusts column widths based on content length, respecting min/max bounds.
    Modifies the Excel file in-place.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to adjust (None = all sheets)
        min_width: Minimum column width (default: 8)
        max_width: Maximum column width (default: 50)
        padding: Extra padding characters (default: 2)
        header_factor: Multiply header length by this factor (default: 1.2)
    
    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet_name not found
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Use defaults if not provided
    config = DEFAULT_AUTO_WIDTH_CONFIG.copy()
    if min_width is not None:
        config['min_width'] = min_width
    if max_width is not None:
        config['max_width'] = max_width
    if padding is not None:
        config['padding'] = padding
    if header_factor is not None:
        config['header_factor'] = header_factor
    
    logger.info(f"Applying auto-width to {excel_path} with config: {config}")
    
    wb = load_workbook(excel_path)
    
    # Determine which sheets to process
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")
        sheets_to_process = [wb[sheet_name]]
    else:
        sheets_to_process = wb.worksheets
    
    # Process each sheet
    for ws in sheets_to_process:
        _adjust_sheet_column_widths(ws, config)
        logger.info(f"Adjusted column widths for sheet: {ws.title}")
    
    wb.save(excel_path)
    logger.info(f"Saved auto-width changes to {excel_path}")


def get_optimal_column_widths(
    excel_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    min_width: Optional[int] = None,
    max_width: Optional[int] = None,
    padding: Optional[int] = None,
    header_factor: Optional[float] = None,
) -> Dict[str, Dict[str, float]]:
    """
    Calculate optimal column widths without modifying the file.
    
    Returns a dictionary mapping sheet names to column width dictionaries.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to analyze (None = all sheets)
        min_width: Minimum column width (default: 8)
        max_width: Maximum column width (default: 50)
        padding: Extra padding characters (default: 2)
        header_factor: Multiply header length by this factor (default: 1.2)
    
    Returns:
        Dictionary: {sheet_name: {column_letter: width, ...}, ...}
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    config = DEFAULT_AUTO_WIDTH_CONFIG.copy()
    if min_width is not None:
        config['min_width'] = min_width
    if max_width is not None:
        config['max_width'] = max_width
    if padding is not None:
        config['padding'] = padding
    if header_factor is not None:
        config['header_factor'] = header_factor
    
    wb = load_workbook(excel_path, read_only=True)
    
    # Determine which sheets to process
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")
        sheets_to_process = [wb[sheet_name]]
    else:
        sheets_to_process = wb.worksheets
    
    result = {}
    for ws in sheets_to_process:
        widths = _calculate_column_widths(ws, config)
        result[ws.title] = widths
    
    wb.close()
    return result


# ============================================================================
# Helper Functions
# ============================================================================

def _adjust_sheet_column_widths(ws, config: Dict) -> None:
    """Adjust column widths for a single worksheet."""
    widths = _calculate_column_widths(ws, config)
    
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _calculate_column_widths(ws, config: Dict) -> Dict[str, float]:
    """Calculate optimal widths for all columns in a worksheet."""
    min_width = config['min_width']
    max_width = config['max_width']
    padding = config['padding']
    header_factor = config['header_factor']
    
    column_widths = {}
    
    # Iterate through columns
    for col_idx, column in enumerate(ws.iter_cols(), start=1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        for idx, cell in enumerate(column):
            if cell.value is not None:
                cell_value = str(cell.value)
                cell_length = len(cell_value)
                
                # Apply header factor to first row
                if idx == 0:
                    cell_length = int(cell_length * header_factor)
                
                max_length = max(max_length, cell_length)
        
        # Apply padding and bounds
        optimal_width = max_length + padding
        optimal_width = max(optimal_width, min_width)
        optimal_width = min(optimal_width, max_width)
        
        column_widths[col_letter] = optimal_width
    
    return column_widths


def apply_auto_width_to_writer(
    writer,
    sheet_name: str,
    min_width: Optional[int] = None,
    max_width: Optional[int] = None,
    padding: Optional[int] = None,
    header_factor: Optional[float] = None,
) -> None:
    """
    Apply auto-width to a sheet in an ExcelWriter object.
    
    Useful when writing DataFrames with pandas ExcelWriter.
    
    Args:
        writer: pandas ExcelWriter object (must use openpyxl engine)
        sheet_name: Name of sheet to adjust
        min_width: Minimum column width (default: 8)
        max_width: Maximum column width (default: 50)
        padding: Extra padding characters (default: 2)
        header_factor: Multiply header length by this factor (default: 1.2)
    
    Example:
        >>> with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        ...     df.to_excel(writer, sheet_name='Sheet1', index=False)
        ...     apply_auto_width_to_writer(writer, 'Sheet1')
    """
    config = DEFAULT_AUTO_WIDTH_CONFIG.copy()
    if min_width is not None:
        config['min_width'] = min_width
    if max_width is not None:
        config['max_width'] = max_width
    if padding is not None:
        config['padding'] = padding
    if header_factor is not None:
        config['header_factor'] = header_factor
    
    # Access the workbook and worksheet
    wb = writer.book
    ws = wb[sheet_name]
    
    _adjust_sheet_column_widths(ws, config)
    logger.info(f"Applied auto-width to sheet '{sheet_name}' in ExcelWriter")


def get_column_widths_from_dataframe(
    df,
    min_width: Optional[int] = None,
    max_width: Optional[int] = None,
    padding: Optional[int] = None,
    header_factor: Optional[float] = None,
) -> Dict[str, float]:
    """
    Calculate optimal column widths for a DataFrame.
    
    Returns a dictionary mapping column letters to widths.
    Useful for pre-calculating widths before writing to Excel.
    
    Args:
        df: pandas DataFrame
        min_width: Minimum column width (default: 8)
        max_width: Maximum column width (default: 50)
        padding: Extra padding characters (default: 2)
        header_factor: Multiply header length by this factor (default: 1.2)
    
    Returns:
        Dictionary: {column_letter: width, ...}
    """
    config = DEFAULT_AUTO_WIDTH_CONFIG.copy()
    if min_width is not None:
        config['min_width'] = min_width
    if max_width is not None:
        config['max_width'] = max_width
    if padding is not None:
        config['padding'] = padding
    if header_factor is not None:
        config['header_factor'] = header_factor
    
    min_width = config['min_width']
    max_width = config['max_width']
    padding = config['padding']
    header_factor = config['header_factor']
    
    column_widths = {}
    
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Header length
        max_length = int(len(str(col_name)) * header_factor)
        
        # Data lengths
        if len(df) > 0:
            col_data = df[col_name].astype(str)
            max_data_length = col_data.str.len().max()
            max_length = max(max_length, max_data_length)
        
        # Apply padding and bounds
        optimal_width = max_length + padding
        optimal_width = max(optimal_width, min_width)
        optimal_width = min(optimal_width, max_width)
        
        column_widths[col_letter] = optimal_width
    
    return column_widths
